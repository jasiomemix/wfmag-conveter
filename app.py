from flask import Flask, render_template, request, send_file, jsonify
import openpyxl
import io
import os
import sys
import zipfile
import threading
from datetime import datetime
from xml.etree.ElementTree import Element, SubElement, ElementTree, indent

FIRMA_NAZWA = 'MOREX GM SPÓŁKA Z OGRANICZONĄ ODPOWIEDZIALNOŚCIĄ'

def resource_path(relative_path):
    if getattr(sys, 'frozen', False):
        base_path = sys._MEIPASS
    else:
        base_path = os.path.dirname(os.path.abspath(__file__))
    return os.path.join(base_path, relative_path)

app = Flask(__name__, template_folder=resource_path('templates'))
app.config['MAX_CONTENT_LENGTH'] = 64 * 1024 * 1024

# ═══════════════════════════════════════════════
# MATRIX 2 (original) — full export with prices
# ═══════════════════════════════════════════════

COL_MAP = {
    'article code':             'symbol',
    'article barcode':          'kod_kreskowy',
    'article name':             'nazwa',
    'gross sell price [pln]':   'cena_sprzedazy_1',
    'net purchase price [pln]': 'cena_zakupu',
}

NOT_SHOP = {
    'article code', 'article barcode', 'article name', 'supplier model code',
    'cat', 'dept', 'grp', 'licence', 'sublicence', 'freestock qty',
    'net purchase price [pln]', 'gross sell price [pln]',
    'order value net purchase price [pln]', 'order value net sell price [pln]',
    'total quantity', 'object no:', 'object no', 'delivery no:', 'order gold no:', '', 'none'
}


def parse_sheet_m2(ws):
    all_rows = list(ws.iter_rows(values_only=True))
    if len(all_rows) < 3:
        return [], {}, {}

    header_row_idx = None
    for i, row in enumerate(all_rows):
        vals = [str(v or '').strip().lower() for v in row]
        if 'article code' in vals or 'article barcode' in vals:
            header_row_idx = i
            break
    if header_row_idx is None:
        return [], {}, {}

    header_row = all_rows[header_row_idx]

    object_nos = {}
    for look_back in range(1, min(4, header_row_idx + 1)):
        candidate = all_rows[header_row_idx - look_back]
        cvals = [str(v or '').strip().lower() for v in candidate]
        if any('object no' in v for v in cvals):
            for i, v in enumerate(candidate):
                if v is not None and str(v).strip().lower() not in ('object no:', 'object no', '', 'none'):
                    try:
                        num = str(int(float(str(v).strip())))
                        object_nos[i] = num
                    except:
                        pass
            break

    col_index = {}
    shop_cols = {}

    for i, cell_val in enumerate(header_row):
        h = str(cell_val or '').strip()
        h_low = h.lower()
        if h_low in COL_MAP:
            col_index[COL_MAP[h_low]] = i
        elif h_low not in NOT_SHOP and h:
            shop_cols[h] = i

    if not shop_cols:
        return [], {}, {}

    shop_to_object_no = {shop: object_nos[idx] for shop, idx in shop_cols.items() if idx in object_nos}

    rows = []
    for row in all_rows[header_row_idx + 1:]:
        if not any(v is not None for v in row):
            continue
        ac_idx = col_index.get('symbol')
        an_idx = col_index.get('nazwa')
        if ac_idx is not None:
            ac = row[ac_idx] if ac_idx < len(row) else None
            an = row[an_idx] if (an_idx is not None and an_idx < len(row)) else None
            if ac is None or str(ac).strip() in ('', '0', 'None'):
                continue
            if an is None or str(an).strip() in ('', 'None'):
                continue

        item = {}
        for field, idx in col_index.items():
            val = row[idx] if idx < len(row) else None
            item[field] = str(val).strip() if val is not None else ''
        for shop, idx in shop_cols.items():
            val = row[idx] if idx < len(row) else None
            try:
                item[f'qty_{shop}'] = int(float(str(val))) if val not in (None, '', 'None') else 0
            except:
                item[f'qty_{shop}'] = 0
        rows.append(item)

    return rows, shop_cols, shop_to_object_no


def parse_excel_m2(file_bytes):
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    all_rows = []
    all_shops_ordered = []
    shop_to_object_no = {}
    shops_seen = set()

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows, shop_cols, s2o = parse_sheet_m2(ws)
        for r in rows:
            r['_sheet'] = sheet_name
        all_rows.extend(rows)
        shop_to_object_no.update(s2o)
        for s in shop_cols:
            if s not in shops_seen:
                all_shops_ordered.append(s)
                shops_seen.add(s)

    return all_rows, all_shops_ordered, shop_to_object_no


# ═══════════════════════════════════════════════
# MATRIX 1 — simple: EAN | Tytuł | shop1 | shop2 | ...
# ═══════════════════════════════════════════════

M1_KNOWN_COLS = {'ean', 'tytuł', 'tytul', 'title', 'nazwa', 'name', 'barcode', 'kod', 'kod kreskowy',
                 'cena', 'cena brutto', 'cena netto', 'cena sprzedaży', 'cena sprzedazy',
                 'cena zakupu', 'price', 'gross price', 'net price', 'cena detal', 'cena katalogowa'}

M1_PRICE_NAMES = {'cena', 'cena brutto', 'cena sprzedaży', 'cena sprzedazy', 'cena detal',
                  'cena katalogowa', 'price', 'gross price', 'cena sprzedazy 1'}
M1_PURCHASE_PRICE_NAMES = {'cena netto', 'cena zakupu', 'net price', 'cena zakupu netto'}


def parse_sheet_m1(ws):
    all_rows = list(ws.iter_rows(values_only=True))
    if len(all_rows) < 2:
        return [], {}, ws.title

    header_row = all_rows[0]
    headers = [str(h or '').strip() for h in header_row]
    headers_low = [h.lower() for h in headers]

    # Find EAN column
    ean_idx = None
    for i, h in enumerate(headers_low):
        if h in ('ean', 'barcode', 'kod', 'kod kreskowy'):
            ean_idx = i
            break
    if ean_idx is None:
        ean_idx = 0

    # Find title/name column
    name_idx = None
    for i, h in enumerate(headers_low):
        if h in ('tytuł', 'tytul', 'title', 'nazwa', 'name'):
            name_idx = i
            break
    if name_idx is None and len(headers) > 1:
        name_idx = 1

    # Find price columns (optional)
    sell_price_idx = None
    purchase_price_idx = None
    for i, h in enumerate(headers_low):
        if sell_price_idx is None and h in M1_PRICE_NAMES:
            sell_price_idx = i
        if purchase_price_idx is None and h in M1_PURCHASE_PRICE_NAMES:
            purchase_price_idx = i

    # Shop columns = everything else (skip ean, name, price cols)
    skip_idxs = {ean_idx, name_idx, sell_price_idx, purchase_price_idx} - {None}
    shop_cols = {}
    for i, h in enumerate(headers):
        if i in skip_idxs:
            continue
        h_low = h.lower()
        if h_low not in M1_KNOWN_COLS and h.strip():
            shop_cols[h] = i

    if not shop_cols:
        return [], {}, ws.title

    rows = []
    for row in all_rows[1:]:
        if not any(v is not None for v in row):
            continue

        ean_val = row[ean_idx] if ean_idx < len(row) else None
        if ean_val is None or str(ean_val).strip() in ('', 'None'):
            continue

        try:
            ean_str = str(int(float(str(ean_val)))).strip()
        except:
            ean_str = str(ean_val).strip()

        name_val = row[name_idx] if (name_idx is not None and name_idx < len(row)) else ''
        name_str = str(name_val).strip() if name_val is not None else ''

        # Read prices if columns exist
        sell_price = '0'
        if sell_price_idx is not None and sell_price_idx < len(row):
            pv = row[sell_price_idx]
            if pv not in (None, '', 'None'):
                try:
                    sell_price = str(round(float(str(pv).replace(',', '.')), 2))
                except:
                    sell_price = '0'

        purchase_price = '0'
        if purchase_price_idx is not None and purchase_price_idx < len(row):
            pv = row[purchase_price_idx]
            if pv not in (None, '', 'None'):
                try:
                    purchase_price = str(round(float(str(pv).replace(',', '.')), 2))
                except:
                    purchase_price = '0'

        item = {
            'kod_kreskowy': ean_str,
            'nazwa': name_str,
            'symbol': '',
            'cena_sprzedazy_1': sell_price,
            'cena_zakupu': purchase_price,
        }

        for shop, idx in shop_cols.items():
            val = row[idx] if idx < len(row) else None
            try:
                item[f'qty_{shop}'] = int(float(str(val))) if val not in (None, '', 'None') else 0
            except:
                item[f'qty_{shop}'] = 0

        rows.append(item)

    return rows, shop_cols, ws.title


def parse_excel_m1(file_bytes):
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    all_rows = []
    all_shops_ordered = []
    shops_seen = set()

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows, shop_cols, _ = parse_sheet_m1(ws)
        for r in rows:
            r['_sheet'] = sheet_name
        all_rows.extend(rows)
        for s in shop_cols:
            if s not in shops_seen:
                all_shops_ordered.append(s)
                shops_seen.add(s)

    return all_rows, all_shops_ordered, {}


# ═══════════════════════════════════════════════
# AUTO-DETECT which matrix
# ═══════════════════════════════════════════════

def detect_matrix(file_bytes):
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i > 10:
                break
            vals = [str(v or '').strip().lower() for v in row]
            if 'article code' in vals or 'article barcode' in vals:
                return 'matrix2'
            if 'ean' in vals or 'tytuł' in vals or 'tytul' in vals:
                return 'matrix1'
    return 'matrix2'


# ═══════════════════════════════════════════════
# XML generator (shared)
# ═══════════════════════════════════════════════

def t(parent, tag, text=''):
    el = SubElement(parent, tag)
    el.text = str(text) if text is not None else ''
    return el


def generate_ecod_xml(rows, firma_nazwa, shop_name, prefix, object_no, seller_iln='', nip='', vat_rate=5, matrix='matrix2'):
    now = datetime.now()
    qty_field = f'qty_{shop_name}'
    shop_rows = [r for r in rows if r.get(qty_field, 0) > 0]
    if not shop_rows:
        return None

    kontrahent_nazwa = f"{prefix} {object_no}".strip() if object_no else (prefix or shop_name)
    order_number = f"{prefix}{object_no}_{now.strftime('%Y%m%d%H%M')}" if object_no else f"ZO_{shop_name}_{now.strftime('%Y%m%d%H%M')}"
    buyer_iln = '5909000850804'
    seller_iln_val = '5909000850804'
    buyer_nip = nip.replace('-', '').replace(' ', '') if nip else ''

    vat_mult = 1 + (vat_rate / 100)

    root = Element('Document-Order')

    hdr = SubElement(root, 'Order-Header')
    t(hdr, 'OrderNumber', order_number)
    t(hdr, 'OrderDate', now.strftime('%Y-%m-%d'))
    t(hdr, 'DocumentFunctionCode', 'O')
    t(hdr, 'Remarks', f'Zamowienie {kontrahent_nazwa}')

    parties = SubElement(root, 'Order-Parties')

    buyer = SubElement(parties, 'Buyer')
    t(buyer, 'ILN', buyer_iln)
    if buyer_nip:
        t(buyer, 'TaxID', buyer_nip)
    t(buyer, 'Name', kontrahent_nazwa)
    t(buyer, 'Country', 'PL')

    seller = SubElement(parties, 'Seller')
    t(seller, 'ILN', seller_iln_val)
    t(seller, 'Name', firma_nazwa)
    t(seller, 'Country', 'PL')

    delivery = SubElement(parties, 'DeliveryPoint')
    t(delivery, 'ILN', buyer_iln)
    t(delivery, 'Name', shop_name)

    order_lines = SubElement(root, 'Order-Lines')

    total_qty = 0
    total_net = 0.0
    total_gross = 0.0

    for i, row in enumerate(shop_rows, start=1):
        barcode_raw = row.get('kod_kreskowy', '')
        try:
            barcode = str(int(float(barcode_raw))).strip() if barcode_raw else ''
        except:
            barcode = str(barcode_raw).strip() if barcode_raw else ''

        nazwa = row.get('nazwa', '')
        symbol = row.get('symbol', '')
        qty = row.get(qty_field, 0)
        # Prices: always read from row data; skip tags only if price is 0 and matrix1
        cena_sp = row.get('cena_sprzedazy_1', '0') or '0'
        try:
            cena_brutto = round(float(str(cena_sp).replace(',', '.')), 2)
            cena_netto = round(cena_brutto / vat_mult, 2)
            net_amount = round(cena_netto * qty, 2)
        except:
            cena_brutto = cena_netto = net_amount = 0.0

        has_price = cena_brutto > 0

        total_qty += qty
        if has_price or matrix == 'matrix2':
            total_net += net_amount
            total_gross += round(cena_brutto * qty, 2)

        line = SubElement(order_lines, 'Line')
        li = SubElement(line, 'Line-Item')
        t(li, 'LineNumber', str(i))
        t(li, 'EAN', barcode if barcode else '0000000000000')
        if symbol:
            t(li, 'BuyerItemCode', symbol)
        t(li, 'ItemDescription', nazwa)
        t(li, 'ItemType', 'CU')
        t(li, 'OrderedQuantity', f'{qty}.000')
        t(li, 'UnitOfMeasure', 'PCE')
        # Include price tags if: matrix2 (always) OR matrix1 with actual price from Excel
        if has_price or matrix == 'matrix2':
            t(li, 'OrderedUnitNetPrice', f'{cena_netto:.2f}')
            t(li, 'OrderedUnitGrossPrice', f'{cena_brutto:.2f}')
            t(li, 'NetAmount', f'{net_amount:.2f}')

    summary = SubElement(root, 'Order-Summary')
    t(summary, 'TotalLines', str(len(shop_rows)))
    t(summary, 'TotalOrderedAmount', f'{total_qty}.000')
    if total_net > 0 or matrix == 'matrix2':
        t(summary, 'TotalNetAmount', f'{total_net:.2f}')
        t(summary, 'TotalGrossAmount', f'{total_gross:.2f}')

    try:
        indent(root, space='  ')
    except AttributeError:
        pass

    tree = ElementTree(root)
    import io as _io
    sbuf = _io.StringIO()
    sbuf.write('<?xml version="1.0" encoding="UTF-8"?>\n')
    tree.write(sbuf, encoding='unicode')
    return sbuf.getvalue().encode('utf-8')


# ═══════════════════════════════════════════════
# ROUTES
# ═══════════════════════════════════════════════

@app.route('/')
def index():
    return render_template('index.html')


@app.route('/detect', methods=['POST'])
def detect():
    if 'file' not in request.files:
        return jsonify({'error': 'Brak pliku'}), 400
    f = request.files['file']
    try:
        file_bytes = f.read()
        matrix_type = detect_matrix(file_bytes)
        return jsonify({'matrix': matrix_type})
    except Exception as e:
        return jsonify({'error': f'Błąd: {str(e)}'}), 400


@app.route('/scan', methods=['POST'])
def scan():
    if 'file' not in request.files:
        return jsonify({'error': 'Brak pliku'}), 400
    f = request.files['file']
    matrix = request.form.get('matrix', 'matrix2')
    try:
        file_bytes = f.read()
        if matrix == 'matrix1':
            rows, shops, shop_to_object_no = parse_excel_m1(file_bytes)
        else:
            rows, shops, shop_to_object_no = parse_excel_m2(file_bytes)

        if not shops:
            return jsonify({'error': 'Nie znaleziono kolumn sklepów. Sprawdź format pliku i wybraną konfigurację.'}), 400
        shop_counts = {s: sum(1 for r in rows if r.get(f'qty_{s}', 0) > 0) for s in shops}
        return jsonify({
            'shops': shops,
            'shop_counts': shop_counts,
            'shop_to_object_no': shop_to_object_no,
            'total_rows': len(rows)
        })
    except Exception as e:
        import traceback
        return jsonify({'error': f'Błąd: {str(e)}', 'detail': traceback.format_exc()}), 400


@app.route('/convert', methods=['POST'])
def convert():
    if 'file' not in request.files:
        return jsonify({'error': 'Brak pliku'}), 400
    f = request.files['file']
    matrix = request.form.get('matrix', 'matrix2')
    firma = FIRMA_NAZWA
    prefix = request.form.get('prefix', '').strip()
    seller_iln = request.form.get('seller_iln', '').strip()
    nip = request.form.get('nip', '').strip()
    selected = request.form.getlist('shops')

    try:
        vat_rate = int(request.form.get('vat_rate', '5'))
    except:
        vat_rate = 5

    if not selected:
        return jsonify({'error': 'Wybierz co najmniej jeden sklep'}), 400

    try:
        file_bytes = f.read()
        if matrix == 'matrix1':
            rows, _, shop_to_object_no = parse_excel_m1(file_bytes)
        else:
            rows, _, shop_to_object_no = parse_excel_m2(file_bytes)
    except Exception as e:
        return jsonify({'error': f'Błąd odczytu: {str(e)}'}), 400

    if len(selected) == 1:
        shop = selected[0]
        obj_no = shop_to_object_no.get(shop, shop)
        content = generate_ecod_xml(rows, firma, shop, prefix, obj_no, seller_iln, nip, vat_rate, matrix)
        if not content:
            return jsonify({'error': f'Brak artykułów z ilością > 0 dla: {shop}'}), 400
        raw = content.encode('utf-8') if isinstance(content, str) else content
        buf = io.BytesIO(raw)
        buf.seek(0)
        safe = f"{prefix}_{obj_no}".strip('_') if prefix else obj_no
        fname = f'ZO_{safe}_{datetime.now().strftime("%Y%m%d")}.xml'
        return send_file(buf, as_attachment=True, download_name=fname, mimetype='application/xml')
    else:
        zip_buf = io.BytesIO()
        with zipfile.ZipFile(zip_buf, 'w', zipfile.ZIP_DEFLATED) as zf:
            ok = 0
            for shop in selected:
                obj_no = shop_to_object_no.get(shop, shop)
                content = generate_ecod_xml(rows, firma, shop, prefix, obj_no, seller_iln, nip, vat_rate, matrix)
                if content:
                    safe = f"{prefix}_{obj_no}".strip('_') if prefix else obj_no
                    fname = f'ZO_{safe}_{datetime.now().strftime("%Y%m%d")}.xml'
                    raw = content.encode('utf-8') if isinstance(content, str) else content
                    zf.writestr(fname, raw)
                    ok += 1
        if ok == 0:
            return jsonify({'error': 'Żaden sklep nie ma artykułów z ilością > 0'}), 400
        zip_buf.seek(0)
        zname = f'zamowienia_wfmag_{datetime.now().strftime("%Y%m%d_%H%M%S")}.zip'
        return send_file(zip_buf, as_attachment=True, download_name=zname, mimetype='application/zip')


def start_flask():
    app.run(host='127.0.0.1', port=5050, debug=False, use_reloader=False)

if __name__ == '__main__':
    if getattr(sys, 'frozen', False):
        import webview
        server_thread = threading.Thread(target=start_flask, daemon=True)
        server_thread.start()
        webview.create_window(
            'WF-Mag Konwerter — MOREX GM',
            'http://127.0.0.1:5050',
            width=800,
            height=900,
            min_size=(600, 500),
        )
        webview.start()
    else:
        app.run(debug=True, port=5050)
